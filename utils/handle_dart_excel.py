from openpyxl import load_workbook
from utils import session_helper
from texts.models import Pupil

def open_dart_excel():
    wb = load_workbook('../DART_test_data.xlsx')
    ws = wb['Blad1']
    lines = list(ws.values)
    header = lines[0]
    data = lines[1:]
    return wb, header, data

def audio_urls(dart_data = None):
    if not dart_data:
        wb, header, dart_data = open_dart_excel()
    output = []
    for line in dart_data:
        audio_url = line[18]
        if audio_url not in output: output.append(audio_url)
    return output


def show_column_unique_values(header,data):
    '''show the number of unique values per column.'''
    for i, cn in enumerate(header):
        print(i,cn, len(set([x[i] for x in data])))

def url_to_session_list(url,data):
    return [x for x in data if x[18] == url]

def session_list_to_word_list(session_list):
    return [get_word(line) for line in session_list]

def session_list_to_correct_list(session_list):
    return [get_correct(line) for line in session_list]

   


def get_all_pupils(data):
    return list(set([get_pupil(line) for line in data]))

def get_all_teachers(data):
    return list(set([get_teacher(line) for line in data]))

def get_all_schools(data):
    return list(set([get_school(line) for line in data]))


def get_word(line):
    return line[2]
def get_correct(line):
    return line[3]

def get_condition(line):
    return line[14]
def get_test_type(line):
    return line[15]
def get_exercise(line):
    return line[17]

def get_list_id(line):
    return line[1]
def get_set_id(line):
    return line[4]
def get_page_id(line):
    return line[16]

def get_pupil(line):
    return line[8]
def get_teacher(line):
    return line[0]
def get_school(line):
    return line[9]

def get_url(line):
    return line[18]
def get_duration(line):
    return line[19]
def get_date_str(line):
    return line[7]


# list set page id makes a word list unique n = 3724
# can use audio filename to select sessions n = 3127

def get_all_list_numbers(dart_data):
    return list(set([x[1] for x in dart_data]))

def get_all_list_set_page_combinations(dart_data):
    list_numbers = get_all_list_numbers(dart_data)
    output = []
    for x in dart_data:
        combination = data_line_to_list_set_page(x)
        if combination not in output: output.append(combination)
    return output

def make_list_page_to_wordlist_dict(dart_data):
    o = {}
    for x in dart_data:
        word = x[2]
        combination = data_line_to_list_set_page(x)
        if combination not in o.keys(): o[combination] = []
        if word not in o[combination]:
            o[combination].append(word)
    return o

def data_line_to_list_set_page(data_line):
    '''helper function to get list, set and page number from a data line.'''
    list_number = data_line[1]
    set_number = data_line[4]
    page_number = data_line[16]
    return list_number, set_number, page_number

def group_info_on_teacher_list_set(info):
    if type(info) == str: info = eval(info)
    d = {}
    for line in info:
        k = line[0],line[1],line[4]
        if k not in d.keys(): d[k] = []
        d[k].append(line)
    keys = list(d.keys())
    for k in keys:
        key = 'words_'+'_'.join(map(str,k))
        d[key] = []
        for line in d[k]:
            d[key].append(line[2])
        # print(key)
        # print(len(d[key]), ' '.join(d[key]))
    return d

def check_group_info_dict(d):
    wordlist_keys = [k for k in d.keys() if 'word' in k]
    number_keys = [k for k in d.keys() if not 'word' in k]
    words_identical = True 
    wordlist = d[wordlist_keys[0]]
    for wordlist_key in wordlist_keys:
        if wordlist != d[wordlist_key]: words_identical = False
    teacher_ids = [k[0] for k in number_keys]
    teachers_differ = len(teacher_ids) == len(set(teacher_ids))
    return words_identical, teachers_differ

def fix_multiple_dart_annotators(session):
    d = group_info_on_teacher_list_set(session.info)
    wi, td =  check_group_info_dict(d)
    if not wi: raise ValueError('words are not identical',session)
    wordlist_keys = [k for k in d.keys() if 'word' in k]
    number_keys = [k for k in d.keys() if not 'word' in k]
    session_list = d[number_keys[0]]
    word_list = d[wordlist_keys[0]]
    if len(wordlist_keys) == 1: raise ValueError('only 1 wl',session)
    if len(session_list) != len(word_list): 
        raise ValueError('wl and sl not of equal length',session)
    cl = session_list_to_correct_list(session_list)
    
    session.multiple_dart_correctors = True
    session.word_list = ','.join(word_list)
    session.nwords = len(word_list)
    session.ncorrect = cl.count(1)
    session.all_correct = cl.count(1) == len(cl)
    session.all_incorrect = cl.count(0) == len(cl)
    session.correct_list = ','.join(map(str,cl))
    session.align = '\n'.join(session_helper.align(session))

    session.save()
    session_helper.to_word_instances(session)


    
def open_pupil_metadata(f = '../metadata_experiment2_anoniem.xlsx'):
    '''load the pupil meta data from the excel file.'''
    wb = load_workbook(f)
    ws = wb['Blad2']
    lines = list(ws.values)
    header = lines[0]
    data = lines[1:]
    return wb, header, data

def make_pupil_id(line):
    '''create the pupil id from the pupil metadata file.'''
    school_id = str(line[2])
    class_id = str(line[3])
    pupil_id = str(line[4])
    pupil_id = school_id + class_id + pupil_id
    pupil_id = pupil_id.lstrip('0')
    assert pupil_id == str(line[0])
    return pupil_id
    
def get_database_pupil(line, verbose = False):
    '''retrieve a pupil from database based on the identifier'''
    pupil_id = make_pupil_id(line)
    try:pupil = Pupil.objects.get(identifier = pupil_id)
    except:
        if verbose: print('could not find',pupil_id, 'skipping')
        return False
    return pupil

def update_pupil_information(line, save = False):
    '''update pupil information based on the meta data in a line from
    the excel metadata file
    '''
    dg = {'jongen':'male','meisje':'female',None:''}
    dll = {'Zon':'zon','Maan':'maan','Ster':'ster',
        'eigen leerlijn': 'eigen leerlijn','onbekend':'onbekend',
        None:'onbekend'}
    p = get_database_pupil(line)
    if not p: return None 
    p.school_id = str(line[2])
    p.class_id = str(line[3])
    p.pupil_id = str(line[4])
    p.birth_date = line[10]
    p.home_lang_str = str(line[13])
    p.gender = dg[line[11]]
    p.reading_level = dll[line[8]]
    p.only_dutch = only_dutch(line)
    p.also_dutch = also_dutch(line)
    p.no_dutch = no_dutch(line)
    p.info = str(line)
    if save: p.save()
    return p

def only_dutch(line):
    if line[13] == None: return None
    l = str(line[13]).lower().strip()
    if l == 'nl': return True
    if l == 'n': return True
    if l == 'nederlands': return True
    if l == 'ned': return True
    return False

def also_dutch(line):
    if line[13] == None: return None
    if only_dutch(line): return False
    l = str(line[13]).lower().strip()
    if 'nederlands' in l: return True
    if 'nederlans' in l: return True
    if 'ned ' in l: return True
    return False

def no_dutch(line):
    if line[13] == None: return None
    if not only_dutch(line) and not also_dutch(line): return True
    return False


def update_all_pupil_information(save = False):
    '''updates pupil information based on pupil meta data'''
    _, header, data = open_pupil_metadata()
    n = 0
    for line in data:
        pupil = update_pupil_information(line, save)
        if pupil: n += 1
    print('updated:',n,'pupils')



    
    
    
        

